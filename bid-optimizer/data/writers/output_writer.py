"""
Output Writer
Writes DataFrames to Excel files with proper formatting
"""

import pandas as pd
from io import BytesIO
from typing import Dict, Optional, Any
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill
from datetime import datetime


class OutputWriter:
    """Writes optimized data to Excel files"""

    # ID columns that should be formatted as text
    ID_COLUMNS = [
        "Campaign ID",
        "Ad Group ID",
        "Portfolio ID",
        "Ad ID",
        "Keyword ID",
        "Product Targeting ID",
    ]

    def __init__(self):
        """Initialize output writer"""
        self.file_buffer = None
        self.sheet_count = 0

    def create_excel_file(self, sheets_data: Dict[str, pd.DataFrame]) -> BytesIO:
        """
        Create Excel file with multiple sheets

        Args:
            sheets_data: Dictionary mapping sheet names to DataFrames
                        e.g., {"Clean Zero Sales": df1, "Working Zero Sales": df2}

        Returns:
            BytesIO buffer containing the Excel file
        """
        # DEBUG: Check what we're getting
        for sheet_name, df in sheets_data.items():
            print(
                f"DEBUG [output_writer]: Sheet '{sheet_name}' has {len(df.columns)} columns"
            )
            print(f"DEBUG [output_writer]: First 5 columns: {list(df.columns[:5])}")
            print(f"DEBUG [output_writer]: Last 5 columns: {list(df.columns[-5:])}")

        # Create BytesIO buffer for output
        output = BytesIO()

        # IMPORTANT: Convert ID columns to string BEFORE writing to Excel
        sheets_data_formatted = {}
        for sheet_name, df in sheets_data.items():
            df_copy = df.copy()

            # Replace NaN values with empty strings for all columns
            df_copy = df_copy.fillna("")

            # Convert ID columns to string to prevent scientific notation
            for col in self.ID_COLUMNS:
                if col in df_copy.columns:
                    # For ID columns, only convert non-empty values to string
                    mask = df_copy[col] != ""
                    df_copy.loc[mask, col] = df_copy.loc[mask, col].astype(str)
                    # Remove any .0 from the end if it exists
                    df_copy.loc[mask, col] = df_copy.loc[mask, col].str.replace(
                        r"\.0$", "", regex=True
                    )

            # Ensure Operation column is set to "Update"
            if "Operation" in df_copy.columns:
                df_copy["Operation"] = "Update"

            sheets_data_formatted[sheet_name] = df_copy

        # Create Excel writer with specific options
        with pd.ExcelWriter(
            output,
            engine="openpyxl",
        ) as writer:
            for sheet_name, df in sheets_data_formatted.items():
                # Write DataFrame to sheet
                df.to_excel(writer, sheet_name=sheet_name, index=False)

                # Get the worksheet for additional formatting
                worksheet = writer.sheets[sheet_name]

                # Apply additional formatting
                self._format_worksheet(worksheet, df)

        # Reset buffer position
        output.seek(0)

        # DEBUG: Check file size
        file_size = len(output.getvalue())
        print(f"DEBUG [output_writer]: Created file with {file_size} bytes")

        return output

    def create_working_file(
        self, optimizations_data: Dict[str, Dict[str, pd.DataFrame]]
    ) -> BytesIO:
        """
        Create Working file with Clean and Working sheets for each optimization

        Args:
            optimizations_data: Dict of optimizations, each containing 'clean' and 'working' DataFrames
                               e.g., {"Zero Sales": {"clean": df1, "working": df2}}

        Returns:
            BytesIO buffer containing the Working Excel file
        """
        sheets_data = {}

        for optimization_name, data in optimizations_data.items():
            # Add Clean sheet
            clean_sheet_name = f"Clean {optimization_name}"
            sheets_data[clean_sheet_name] = data["clean"].copy()

            # Add Working sheet
            working_sheet_name = f"Working {optimization_name}"
            sheets_data[working_sheet_name] = data["working"].copy()

        return self.create_excel_file(sheets_data)

    def create_clean_file(self, optimizations_data: Dict[str, pd.DataFrame]) -> BytesIO:
        """
        Create Clean file with only Clean sheets

        Args:
            optimizations_data: Dict mapping optimization names to DataFrames
                               e.g., {"Zero Sales": df1, "Portfolio Bid": df2}

        Returns:
            BytesIO buffer containing the Clean Excel file
        """
        sheets_data = {}

        for optimization_name, df in optimizations_data.items():
            sheet_name = f"Clean {optimization_name}"
            sheets_data[sheet_name] = df.copy()

        return self.create_excel_file(sheets_data)

    def _format_worksheet(self, worksheet, df):
        """
        Apply formatting to worksheet

        Args:
            worksheet: Openpyxl worksheet object
            df: DataFrame for column width calculation
        """
        # Set column widths based on content
        for column in worksheet.columns:
            max_length = 0
            column_letter = column[0].column_letter

            for cell in column:
                try:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                except:
                    pass

            # Set width (min 10, max 50)
            adjusted_width = min(max(max_length + 2, 10), 50)
            worksheet.column_dimensions[column_letter].width = adjusted_width

        # Format header row
        header_font = Font(bold=True)
        header_fill = PatternFill(
            start_color="F0F0F0", end_color="F0F0F0", fill_type="solid"
        )

        for cell in worksheet[1]:  # First row
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")

        # Freeze top row
        worksheet.freeze_panes = "A2"

        # Additional formatting for ID columns - set cell format to text
        for col_idx, col_name in enumerate(df.columns, 1):
            if col_name in self.ID_COLUMNS:
                column_letter = openpyxl.utils.get_column_letter(col_idx)
                # Set entire column to text format
                for row in range(1, len(df) + 2):  # +2 because row 1 is header
                    cell = worksheet[f"{column_letter}{row}"]
                    cell.number_format = "@"  # Text format

        # Format number columns
        number_columns = [
            "Bid",
            "Daily Budget",
            "Ad Group Default Bid",
            "Spend",
            "Sales",
            "CPC",
            "ROAS",
            "ACOS",
        ]

        for col_idx, col_name in enumerate(df.columns, 1):
            if col_name in number_columns:
                column_letter = openpyxl.utils.get_column_letter(col_idx)
                for row in range(2, len(df) + 2):
                    cell = worksheet[f"{column_letter}{row}"]
                    if cell.value and isinstance(cell.value, (int, float)):
                        if col_name in ["Bid", "CPC"]:
                            cell.number_format = "0.000"  # 3 decimal places
                        else:
                            cell.number_format = "0.00"  # 2 decimal places

    def add_metadata_sheet(self, workbook, metadata: Dict[str, Any]):
        """
        Add metadata sheet with processing information

        Args:
            workbook: Openpyxl workbook object
            metadata: Dictionary with processing metadata
        """
        # Create metadata sheet
        ws = workbook.create_sheet("_Metadata", 0)

        # Add metadata
        ws["A1"] = "Processing Information"
        ws["A1"].font = Font(bold=True, size=14)

        row = 3
        for key, value in metadata.items():
            ws[f"A{row}"] = key
            ws[f"B{row}"] = str(value)
            row += 1

        # Adjust column widths
        ws.column_dimensions["A"].width = 25
        ws.column_dimensions["B"].width = 40

    def get_file_stats(self, file_buffer: BytesIO) -> Dict[str, Any]:
        """
        Get statistics about the generated file

        Args:
            file_buffer: BytesIO buffer containing the file

        Returns:
            Dictionary with file statistics
        """
        # Get file size
        file_size = len(file_buffer.getvalue())

        # Read file to get sheet info
        file_buffer.seek(0)
        excel_file = pd.ExcelFile(file_buffer)
        sheet_names = excel_file.sheet_names

        # Count total rows
        total_rows = 0
        for sheet in sheet_names:
            df = pd.read_excel(file_buffer, sheet_name=sheet)
            total_rows += len(df)

        # Reset buffer position
        file_buffer.seek(0)

        return {
            "size_bytes": file_size,
            "size_mb": round(file_size / (1024 * 1024), 2),
            "sheet_count": len(sheet_names),
            "sheet_names": sheet_names,
            "total_rows": total_rows,
            "generated_at": datetime.now().isoformat(),
        }
